"""
interaccion_loader.py — Carga dinámica de frases conversacionales shipibo↔español.

Lee actions/corpus/frases_conversacionales.xlsx y expone funciones para:
  • Buscar una frase por su forma en español o shipibo
  • Devolver la equivalencia en el idioma contrario
  • Listar frases ejemplo para mostrar al usuario

Filtra automáticamente las frases que NO tengan validado_asesor='SI', así el
sistema solo usa material aprobado por el especialista lingüístico.

Estructura del Excel:
  Columna A: id
  Columna B: categoria        (saludo, agradecer, ayuda, cortesia, etc.)
  Columna C: es               (frase en español)
  Columna D: shp              (frase en shipibo)
  Columna E: tipo             (informal, formal, pregunta, etc.)
  Columna F: fuente           (corpus o 'propuesta (revisar)')
  Columna G: validado_asesor  ('SI' para activar, vacío o cualquier otra cosa = inactiva)
  Columna H: notas_asesor     (libre, para uso del asesor)
"""

import os
import unicodedata
import logging
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)

# ── Ruta al corpus de frases ──────────────────────────────────────────────────
FRASES_PATH = os.path.join(
    os.path.dirname(__file__), "corpus", "frases_conversacionales.xlsx"
)

# Si True, ignora la validación del asesor y carga todas las frases.
# Útil para desarrollo cuando el asesor aún no entregó las validaciones.
# En producción debe estar en False para que solo se usen frases validadas.
MODO_DESARROLLO = False


def _normalizar(texto: Any) -> str:
    """Minúsculas, sin tildes, sin puntuación, sin espacios extras."""
    if not texto:
        return ""
    t = str(texto).lower().strip()
    for c in ".,!?¿¡;:":
        t = t.replace(c, "")
    t = "".join(
        ch for ch in unicodedata.normalize("NFD", t)
        if unicodedata.category(ch) != "Mn"
    )
    return " ".join(t.split())  # colapsa espacios múltiples


def _cargar_desde_excel(path: str) -> List[Dict[str, Any]]:
    """Lee el Excel y devuelve una lista de frases (cada una como dict)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl no instalado. Ejecutá: pip install openpyxl")

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"No se encontró frases_conversacionales.xlsx en '{path}'.\n"
            "Copiá el archivo a actions/corpus/frases_conversacionales.xlsx"
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["frases"] if "frases" in wb.sheetnames else wb.active

    frases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        try:
            id_         = str(row[0]).strip()
            categoria   = str(row[1]).strip() if row[1] else ""
            es          = str(row[2]).strip() if row[2] else ""
            shp         = str(row[3]).strip() if row[3] else ""
            tipo        = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            fuente      = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            validado    = str(row[6]).strip().upper() if len(row) > 6 and row[6] else ""
            notas       = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        except (IndexError, ValueError):
            continue

        if not es or not shp or not categoria:
            continue

        # Filtro de validación: solo frases aprobadas por el asesor.
        if not MODO_DESARROLLO and validado != "SI":
            continue

        frases.append({
            "id":         id_,
            "categoria":  categoria,
            "es":         es,
            "shp":        shp,
            "tipo":       tipo,
            "fuente":     fuente,
            "validado":   validado == "SI",
            "notas":      notas,
            "es_norm":    _normalizar(es),
            "shp_norm":   _normalizar(shp),
        })

    wb.close()
    return frases


def _construir_indices(frases: List[Dict[str, Any]]):
    """
    Devuelve dos índices:
      • por_es:  texto_normalizado → frase
      • por_shp: texto_normalizado → frase
    También índice por categoría para listar opciones.
    """
    por_es: Dict[str, Dict[str, Any]] = {}
    por_shp: Dict[str, Dict[str, Any]] = {}
    por_categoria: Dict[str, List[Dict[str, Any]]] = {}

    for f in frases:
        por_es[f["es_norm"]] = f
        por_shp[f["shp_norm"]] = f
        por_categoria.setdefault(f["categoria"], []).append(f)

    return por_es, por_shp, por_categoria


# ── Carga única al importar ───────────────────────────────────────────────────

try:
    FRASES: List[Dict[str, Any]] = _cargar_desde_excel(FRASES_PATH)
    POR_ES, POR_SHP, POR_CATEGORIA = _construir_indices(FRASES)
    _carga_ok = True
    logger.info(
        "interaccion_loader: %d frases cargadas (modo_desarrollo=%s)",
        len(FRASES), MODO_DESARROLLO
    )
except Exception as e:
    logger.error("interaccion_loader: error al cargar frases. %s", e)
    FRASES = []
    POR_ES = POR_SHP = {}
    POR_CATEGORIA = {}
    _carga_ok = False


# ── API pública ───────────────────────────────────────────────────────────────

def frases_disponibles() -> bool:
    """True si al menos una frase fue cargada correctamente."""
    return _carga_ok and len(FRASES) > 0


def buscar_frase(texto: str) -> Optional[Dict[str, Any]]:
    """
    Busca una frase por su forma en español o shipibo (exacta, normalizada).
    Devuelve la frase completa con un campo extra `idioma_detectado` que
    indica si el match fue 'es' o 'shp'. Devuelve None si no hay match.
    """
    t = _normalizar(texto)
    if not t:
        return None

    if t in POR_ES:
        return {**POR_ES[t], "idioma_detectado": "es"}
    if t in POR_SHP:
        return {**POR_SHP[t], "idioma_detectado": "shp"}

    # Búsqueda por contención: si el texto del usuario contiene una frase
    # corta del corpus (1-2 tokens), también es match.
    tokens = set(t.split())
    for f_norm, f in POR_ES.items():
        if " " not in f_norm and f_norm in tokens:
            return {**f, "idioma_detectado": "es"}
    for f_norm, f in POR_SHP.items():
        if " " not in f_norm and f_norm in tokens:
            return {**f, "idioma_detectado": "shp"}

    return None


def frases_por_categoria(categoria: str) -> List[Dict[str, Any]]:
    """Devuelve todas las frases de una categoría dada."""
    return POR_CATEGORIA.get(categoria, [])


def categorias_disponibles() -> List[str]:
    """Devuelve las categorías que tienen al menos una frase cargada."""
    return list(POR_CATEGORIA.keys())


def frases_ejemplo(n: int = 4) -> List[Dict[str, Any]]:
    """
    Devuelve hasta N frases representativas, una por categoría si es posible,
    para mostrar como sugerencias al usuario al entrar al modo conversación.
    """
    ejemplos = []
    for cat in ["saludo", "agradecer", "cortesia", "ayuda"]:
        if cat in POR_CATEGORIA and POR_CATEGORIA[cat]:
            ejemplos.append(POR_CATEGORIA[cat][0])
            if len(ejemplos) >= n:
                break
    # Si quedó corto, completar con cualquier categoría
    if len(ejemplos) < n:
        for cat, lista in POR_CATEGORIA.items():
            if lista and lista[0] not in ejemplos:
                ejemplos.append(lista[0])
                if len(ejemplos) >= n:
                    break
    return ejemplos[:n]


def total_frases() -> int:
    """Cuenta total de frases cargadas."""
    return len(FRASES)


def total_validadas() -> int:
    """Cuántas frases tienen validado_asesor = 'SI'."""
    return sum(1 for f in FRASES if f["validado"])
