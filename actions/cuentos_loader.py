"""
cuentos_loader.py — Carga dinámica de cuentos interactivos shipibo-konibo.

Reemplaza el cuento embebido en actions.py por una fuente de datos
externa (Excel curado y validado por especialista lingüístico).

Estructura esperada del Excel (cuentos.xlsx):
  Columna A: cuento_id            (ej. pescador_shipibo)
  Columna B: cuento_titulo        (ej. El pescador shipibo)
  Columna C: orden                (número entero, 1, 2, 3, ...)
  Columna D: texto                (párrafo del fragmento)
  Columna E: pregunta             (vacío si no hay pregunta)
  Columna F: respuesta_esperada   (palabra shipibo canónica, vacío si no aplica)
  Columna G: ayuda                (pista pedagógica, vacío si no aplica)

Ubicación del Excel:
  actions/corpus/cuentos.xlsx

Uso en actions.py:
  from cuentos_loader import CUENTOS, cuento_por_id, fragmento, total_fragmentos
"""

import os
import logging
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)

# ── Ruta al corpus de cuentos ─────────────────────────────────────────────────
CUENTOS_PATH = os.path.join(os.path.dirname(__file__), "corpus", "cuentos.xlsx")

# Cuento que se ofrece por defecto cuando el usuario solo dice "quiero un cuento".
# Debe ser un cuento con buena cobertura de vocabulario (todas sus palabras meta
# enseñadas en palabras.xlsx) para que la primera experiencia no sea la más difícil.
CUENTO_PREDETERMINADO = "anciano_camungo"


def _normalizar(s: Any) -> str:
    """Quita espacios y maneja None."""
    if s is None:
        return ""
    return str(s).strip()


def _cargar_desde_excel(path: str) -> Dict[str, Dict[str, Any]]:
    """
    Lee el Excel y devuelve:
    {
      "anciano_camungo": {
        "id":      "anciano_camungo",
        "titulo":  "El anciano y el camungo",
        "fragmentos": [
          {"orden": 1, "texto": "...", "pregunta": "...",
           "respuesta_esperada": "yomerati", "ayuda": "..."},
          ...
        ]
      },
      "otro_cuento": {...}
    }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecutá: pip install openpyxl")

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"No se encontró cuentos en '{path}'.\n"
            "Copiá cuentos.xlsx a actions/corpus/cuentos.xlsx"
        )

    wb = load_workbook(path, read_only=True, data_only=True)

    # Buscar la hoja "cuentos" o la primera si no existe
    if "cuentos" in wb.sheetnames:
        ws = wb["cuentos"]
    else:
        ws = wb.active

    cuentos: Dict[str, Dict[str, Any]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        cuento_id     = _normalizar(row[0])
        cuento_titulo = _normalizar(row[1])
        orden_raw     = row[2]
        texto         = _normalizar(row[3])

        if not cuento_id or not texto or orden_raw is None:
            continue

        try:
            orden = int(orden_raw)
        except (ValueError, TypeError):
            continue

        pregunta           = _normalizar(row[4]) if len(row) > 4 else ""
        respuesta_esperada = _normalizar(row[5]) if len(row) > 5 else ""
        ayuda              = _normalizar(row[6]) if len(row) > 6 else ""

        fragmento_data = {
            "orden":              orden,
            "texto":              texto,
            "pregunta":           pregunta if pregunta else None,
            "respuesta_esperada": respuesta_esperada if respuesta_esperada else None,
            "ayuda":              ayuda if ayuda else None,
        }

        if cuento_id not in cuentos:
            cuentos[cuento_id] = {
                "id":         cuento_id,
                "titulo":     cuento_titulo or cuento_id,
                "fragmentos": [],
            }

        cuentos[cuento_id]["fragmentos"].append(fragmento_data)

    # Ordenar fragmentos por su campo "orden" dentro de cada cuento
    for c in cuentos.values():
        c["fragmentos"].sort(key=lambda f: f["orden"])

    wb.close()
    return cuentos


# ── Carga única al importar el módulo ─────────────────────────────────────────

try:
    CUENTOS: Dict[str, Dict[str, Any]] = _cargar_desde_excel(CUENTOS_PATH)
    _carga_ok = True
    logger.info("cuentos_loader: %d cuento(s) cargado(s)", len(CUENTOS))
except Exception as _e:
    logger.error("cuentos_loader: no se pudo cargar el Excel. Error: %s", _e)
    CUENTOS = {}
    _carga_ok = False


# ── API pública ───────────────────────────────────────────────────────────────

def cuentos_disponibles() -> bool:
    """True si al menos un cuento se cargó correctamente."""
    return _carga_ok and len(CUENTOS) > 0


def lista_cuentos() -> List[Dict[str, str]]:
    """
    Devuelve metadatos de todos los cuentos disponibles:
    [{"id": "pescador_shipibo", "titulo": "El pescador shipibo", "fragmentos": 4}, ...]
    """
    return [
        {"id": c["id"], "titulo": c["titulo"], "fragmentos": len(c["fragmentos"])}
        for c in CUENTOS.values()
    ]


def cuento_por_id(cuento_id: str) -> Optional[Dict[str, Any]]:
    """Devuelve el cuento completo por su id, o None si no existe."""
    return CUENTOS.get(cuento_id)


def total_fragmentos(cuento_id: str) -> int:
    """Cantidad de fragmentos del cuento. 0 si no existe."""
    c = CUENTOS.get(cuento_id)
    return len(c["fragmentos"]) if c else 0


def fragmento(cuento_id: str, indice: int) -> Optional[Dict[str, Any]]:
    """
    Devuelve el fragmento en la posición `indice` (0-based) del cuento.
    Devuelve None si el cuento no existe o el índice está fuera de rango.
    """
    c = CUENTOS.get(cuento_id)
    if not c:
        return None
    fragmentos = c["fragmentos"]
    if 0 <= indice < len(fragmentos):
        return fragmentos[indice]
    return None


def titulo_de(cuento_id: str) -> str:
    """Devuelve el título legible del cuento, o el id como fallback."""
    c = CUENTOS.get(cuento_id)
    return c["titulo"] if c else cuento_id