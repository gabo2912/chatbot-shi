"""
corpus_loader.py — Carga dinámica del corpus léxico shipibo-konibo.

Lee actions/corpus/palabras.xlsx y expone:
  • VOCABULARIO         dict por categoría con las palabras válidas
  • DICCIONARIO         índice bidireccional es↔shp para traducción
  • TOTAL_POR_CATEGORIA dict dinámico {categoria: cantidad} (Capa: corpus dinámico)
  • categorias()        lista de categorías presentes
  • encontrar_palabra() búsqueda por es dentro de categoría
  • siguiente_palabra() rotación al siguiente
  • traducir()          traducción rápida usando el diccionario

Estructura esperada del Excel (palabras.xlsx):
  Columna A: id propuesto    (ej. nat_001, num_001, ali_001 …)
  Columna B: es              (palabra en español)
  Columna C: shp             (forma canónica shipibo o "no hay")
  Columna D: Sinonimos       (variantes aceptables, opcional)
  Columna E: Observaciones   (nota libre para revisión humana)

POLÍTICAS DE FILTRADO:
  • Filas con shp = "no hay" se filtran automáticamente (no entran al sistema).
  • Filas con prefijo desconocido en CATEGORIA_MAP se ignoran silenciosamente
    (permite agregar nuevas categorías SOLO actualizando este mapa).
  • Notas entre paréntesis se extraen: "yapa (pescado vivo)" → canónica "yapa".
"""

import os
import re
import unicodedata
import logging
from typing import Dict, List, Any

logger = logging.getLogger(__name__)

# ── Ruta al corpus ────────────────────────────────────────────────────────────
CORPUS_PATH = os.path.join(os.path.dirname(__file__), "corpus", "palabras.xlsx")

# ── Mapeo prefijo → nombre de categoría ──────────────────────────────────────
# Para agregar una categoría nueva, basta con sumar una entrada a este dict
# Y actualizar los emojis abajo, el slot en domain.yml y el grid del frontend.
CATEGORIA_MAP: Dict[str, str] = {
    "nat":  "naturaleza",
    "ani":  "animales",
    "cuer": "cuerpo",
    "col":  "colores",
    "obj":  "objetos",
    "num":  "números",
    "per":  "personas",
}

# ── Valores que marcan ausencia de equivalente shipibo ────────────────────────
_MARCADORES_SIN_SHP = {"no hay", "no_hay", "n/a", "-", ""}

# ── Pistas pedagógicas por palabra en español ─────────────────────────────────
PISTAS: Dict[str, str] = {
    # naturaleza
    "agua":      "Sin esto no podemos vivir. Corre en los ríos de la Amazonía.",
    "sol":       "Aparece en el cielo durante el día y nos da luz y calor.",
    "árbol":     "Crece alto en la selva. Tiene tronco, ramas y hojas.",
    "luna":      "La vemos brillar en el cielo por las noches.",
    "río":       "Corriente de agua donde los shipibos navegan y pescan.",
    "viento":    "Lo sentimos cuando el aire se mueve entre los árboles.",
    "lluvia":    "Cae del cielo y hace crecer las plantas de la selva.",
    "nube":      "Flota en el cielo y a veces trae lluvia.",
    "fuego":     "Se usa para cocinar. Produce calor y luz.",
    "tierra":    "El suelo sobre el que caminamos y donde crecen las plantas.",
    "noche":     "El momento en que el sol se va y aparecen las estrellas.",
    # animales
    "pez":       "Animal que vive en el agua y se pesca en el río.",
    "tortuga":   "Tiene caparazón duro. Vive en ríos y en la tierra.",
    "perro":     "Animal doméstico que ladra y acompaña a las familias.",
    "mono":      "Trepa los árboles con agilidad. Muy común en la selva.",
    "ave":       "Tiene plumas y puede volar por el cielo.",
    "serpiente": "Reptil sin patas que se arrastra por el suelo.",
    "jaguar":    "El felino más grande de la Amazonía. Muy respetado.",
    "caimán":    "Reptil grande que vive en los ríos amazónicos.",
    "pato":      "Ave acuática que nada en ríos y lagunas.",
    "gallina":   "Ave de corral que pone huevos.",
    "mariposa":  "Insecto con alas de colores que vuela entre las flores.",
    "venado":    "Animal de cuatro patas con cuernos que vive en el bosque.",
    "gato":      "Animal doméstico pequeño que maúlla.",
    # cuerpo
    "mano":      "La usamos para escribir, agarrar y crear cosas.",
    "boca":      "Sirve para comer, hablar y sonreír.",
    "ojo":       "Con esto vemos el mundo que nos rodea.",
    "cabeza":    "La parte más alta de nuestro cuerpo.",
    "pie":       "La usamos para caminar y pararnos.",
    "corazón":   "Órgano que late dentro del pecho y bombea sangre.",
    "cabello":   "Crece en la cabeza. Puede ser largo o corto.",
    "diente":    "Lo usamos para masticar la comida.",
    "nariz":     "Con esto respiramos y olemos.",
    "oreja":     "Con esto escuchamos los sonidos.",
    "brazo":     "Une la mano con el hombro.",
    "pierna":    "La usamos para caminar y correr.",
    # colores (gris/marrón/rosado/morado quedaron filtrados: "no hay" en shp)
    "blanco":    "El color de las nubes y la leche.",
    "negro":     "El color de la oscuridad de la noche.",
    "azul":      "El color del cielo despejado y del agua profunda.",
    "rojo":      "El color de la sangre y del fuego.",
    "verde":     "El color de las plantas y la selva amazónica.",
    "amarillo":  "El color del sol y del oro.",
    # objetos
    "canoa":     "Embarcación de madera que se usa para navegar en el río.",
    "casa":      "Lugar donde vive la familia shipibo.",
    "olla":      "Recipiente de barro que se usa para cocinar.",
    "tela":      "Material tejido que se usa para hacer ropa y artesanías.",
    "comida":    "Todo lo que comemos para vivir y crecer.",
    "ropa":      "Prendas que usamos para vestirnos.",
    "plato":     "Recipiente donde ponemos la comida para servirla.",
    "libro":     "Tiene páginas y contiene conocimiento escrito.",
    "escuela":   "Lugar donde vamos a aprender y estudiar.",
    # números (nueva categoría)
    "uno":       "El primer número, el comienzo de todo conteo.",
    "dos":       "Sigue al uno. Cantidad de manos que tenemos.",
    "tres":      "Cantidad de comidas principales del día.",
    "cuatro":    "Cantidad de patas de un perro o un jaguar.",
    "cinco":     "Cantidad de dedos en una mano.",
    "seis":      "Sigue al cinco. Cantidad de lados de un hexágono.",
    "siete":     "Cantidad de días de la semana.",
    "ocho":      "Cantidad de patas de una araña.",
    "nueve":     "Sigue al ocho. Antes del diez.",
    "diez":      "Cantidad de dedos en las dos manos.",
}


# ── Helpers internos ──────────────────────────────────────────────────────────

def _normalizar(texto: str) -> str:
    """Minúsculas, sin tildes, sin puntuación, sin espacios extra."""
    if not texto:
        return ""
    t = str(texto).lower().strip()
    for c in ".,!?¿¡;:":
        t = t.replace(c, "")
    t = "".join(
        ch for ch in unicodedata.normalize("NFD", t)
        if unicodedata.category(ch) != "Mn"
    )
    return " ".join(t.split())


def _extraer_canonica(texto: str) -> str:
    """
    Extrae la palabra antes del paréntesis aclaratorio.
    "yapa (pescado vivo)" → "yapa"
    "piti (pescado cocinado)" → "piti"
    "mishito" → "mishito"
    Conserva mayúsculas/minúsculas; no normaliza.
    """
    if not texto:
        return ""
    s = str(texto).strip()
    if "(" in s:
        s = s.split("(", 1)[0].strip()
    return s


def _es_marcador_vacio(texto: str) -> bool:
    """True si el texto representa 'no hay equivalente'."""
    if not texto:
        return True
    return str(texto).strip().lower() in _MARCADORES_SIN_SHP


def _parsear_sinonimos(texto: str) -> List[str]:
    """
    Devuelve lista de variantes normalizadas desde la columna 'Sinonimos'.
    Soporta separadores: coma, punto y coma, pipe. Cada entrada puede tener
    un paréntesis aclaratorio que se descarta para el matching.

    Ejemplos:
      "kabori (tortuga de agua)"           → ["kabori"]
      "piti (pescado cocinado), pira"       → ["piti", "pira"]
      "mishito"                              → ["mishito"]
      "no hay"                               → []
    """
    if _es_marcador_vacio(texto):
        return []
    variantes = []
    partes = re.split(r"[,;|]", str(texto))
    for parte in partes:
        canonica = _extraer_canonica(parte)
        v = _normalizar(canonica)
        if v and len(v) >= 2:
            variantes.append(v)
    return variantes


def _safe_cell(row, idx: int) -> str:
    """Acceso seguro a celda; devuelve string vacío si está fuera de rango."""
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


# ── Carga principal ───────────────────────────────────────────────────────────

def _cargar_desde_excel(path: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Lee el Excel y devuelve un dict por categoría.
    Cada palabra es un dict con keys: id, es, shp, shp_display, variantes, pista.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl no instalado. Ejecutá: pip install openpyxl")

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"No se encontró el corpus en '{path}'. "
            "Copiá palabras.xlsx a actions/corpus/palabras.xlsx"
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    vocab: Dict[str, List[Dict[str, Any]]] = {
        nombre: [] for nombre in CATEGORIA_MAP.values()
    }

    # Estadísticas para log
    descartadas_sin_shp = 0
    descartadas_sin_es = 0
    descartadas_prefijo = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        id_p   = _safe_cell(row, 0)
        es     = _safe_cell(row, 1)
        shp_raw = _safe_cell(row, 2)
        sinonimos_raw = _safe_cell(row, 3)

        if not id_p:
            continue

        # Filtrar filas sin español o con shp vacío/no-hay
        if _es_marcador_vacio(es):
            descartadas_sin_es += 1
            continue
        if _es_marcador_vacio(shp_raw):
            descartadas_sin_shp += 1
            continue

        # Resolver categoría desde el prefijo
        prefix = id_p.split("_", 1)[0] if "_" in id_p else ""
        categoria = CATEGORIA_MAP.get(prefix)
        if not categoria:
            descartadas_prefijo += 1
            continue

        # Extraer forma canónica para matching (sin paréntesis) y forma display
        shp_canonica = _extraer_canonica(shp_raw)
        shp_norm = _normalizar(shp_canonica)

        # Variantes desde la columna Sinonimos
        variantes = _parsear_sinonimos(sinonimos_raw)

        # Si el shp original tenía paréntesis, agregar la forma completa como
        # variante aceptable (por si el usuario tipea la versión completa).
        if shp_raw != shp_canonica:
            variantes.append(_normalizar(shp_raw))

        # Deduplicar y quitar la forma canónica si quedó incluida
        variantes = list(dict.fromkeys(
            v for v in variantes if v and v != shp_norm
        ))

        entrada: Dict[str, Any] = {
            "id":          id_p,
            "es":          es,
            "shp":         shp_canonica,    # forma para matching
            "shp_display": shp_raw,         # forma con nota (si la tiene)
            "variantes":   variantes,
            "pista":       PISTAS.get(es.lower(),
                                      "Piensa en el contexto de esta palabra."),
        }
        vocab[categoria].append(entrada)

    wb.close()

    if descartadas_sin_shp or descartadas_sin_es or descartadas_prefijo:
        logger.info(
            "corpus_loader: descartadas %d sin shp, %d sin es, %d prefijo desconocido",
            descartadas_sin_shp, descartadas_sin_es, descartadas_prefijo
        )

    return vocab


def _construir_diccionario(
    vocab: Dict[str, List[Dict[str, Any]]]
) -> Dict[str, str]:
    """Diccionario plano bidireccional para traducción rápida es↔shp."""
    d: Dict[str, str] = {}
    for palabras in vocab.values():
        for p in palabras:
            es_n  = _normalizar(p["es"])
            shp_n = _normalizar(p["shp"])
            d[es_n]  = p["shp"]
            d[shp_n] = p["es"]
            for v in p.get("variantes", []):
                if v and v not in d:
                    d[v] = p["es"]
    return d


# ── Carga única al importar el módulo ────────────────────────────────────────

try:
    VOCABULARIO: Dict[str, List[Dict[str, Any]]] = _cargar_desde_excel(CORPUS_PATH)
    DICCIONARIO: Dict[str, str] = _construir_diccionario(VOCABULARIO)
    _carga_ok = True
    _total_palabras = sum(len(p) for p in VOCABULARIO.values())
    logger.info(
        "corpus_loader: %d palabras cargadas en %d categorías",
        _total_palabras, len([c for c, p in VOCABULARIO.items() if p])
    )
except Exception as _e:
    logger.error("corpus_loader: error al cargar corpus. %s", _e)
    VOCABULARIO = {nombre: [] for nombre in CATEGORIA_MAP.values()}
    DICCIONARIO = {}
    _carga_ok = False


# ── API pública ───────────────────────────────────────────────────────────────

def corpus_disponible() -> bool:
    """True si el corpus se cargó correctamente desde el Excel."""
    return _carga_ok


def categorias() -> List[str]:
    """Devuelve nombres de categorías que tienen al menos una palabra cargada."""
    return [c for c, palabras in VOCABULARIO.items() if palabras]


def palabras_de(categoria: str) -> List[Dict[str, Any]]:
    """Devuelve todas las palabras de una categoría."""
    return VOCABULARIO.get(categoria, [])


def total_por_categoria() -> Dict[str, int]:
    """
    Cuenta dinámica de palabras por categoría (computada desde el corpus
    cargado). Reemplaza la constante TOTAL_PALABRAS hardcoded que estaba
    repetida en db.py, server.py e index.html.
    """
    return {cat: len(palabras) for cat, palabras in VOCABULARIO.items() if palabras}


def total_palabras() -> int:
    """Total de palabras válidas en el corpus completo."""
    return sum(len(p) for p in VOCABULARIO.values())


def encontrar_palabra(categoria: str, es: str) -> Dict[str, Any]:
    """Busca una palabra por su forma en español dentro de una categoría."""
    for p in VOCABULARIO.get(categoria, []):
        if p["es"] == es:
            return p
    return {}


def siguiente_palabra(categoria: str, palabra_actual: str) -> Dict[str, Any]:
    """Devuelve la siguiente palabra en la categoría (rota al inicio)."""
    palabras = VOCABULARIO.get(categoria, [])
    if not palabras:
        return {}
    for i, p in enumerate(palabras):
        if p["es"] == palabra_actual:
            return palabras[(i + 1) % len(palabras)]
    return palabras[0]


def traducir(palabra: str) -> str:
    """Traducción rápida ES⇄SHP usando el diccionario plano."""
    return DICCIONARIO.get(_normalizar(palabra), "")