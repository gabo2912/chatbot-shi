#!/usr/bin/env python3
"""
diagnostico.py — Verifica que todos los archivos del MVP estén en la
ubicación correcta y con el contenido actualizado.

Uso (desde la raíz del proyecto chatbot-shi):
    python3 diagnostico.py
"""

import os
import sys
import json
import sqlite3
from pathlib import Path

# ANSI colors para output legible
class C:
    OK = '\033[92m'
    WARN = '\033[93m'
    ERR = '\033[91m'
    END = '\033[0m'
    BOLD = '\033[1m'

def ok(msg):   print(f"  {C.OK}✓{C.END} {msg}")
def warn(msg): print(f"  {C.WARN}⚠{C.END} {msg}")
def err(msg):  print(f"  {C.ERR}✗{C.END} {msg}")
def head(msg): print(f"\n{C.BOLD}── {msg} ──{C.END}")

HERE = Path.cwd()
errores = 0

print(f"{C.BOLD}═══════════════════════════════════════════════════════════════")
print(f"  DIAGNÓSTICO PISHICO BOT")
print(f"  Directorio actual: {HERE}")
print(f"═══════════════════════════════════════════════════════════════{C.END}")

# ── 1. ARCHIVOS PRESENTES ────────────────────────────────────────────────
head("1. Archivos en la ubicación correcta")
archivos = {
    "actions/actions.py":              True,
    "actions/corpus_loader.py":        True,
    "actions/db.py":                   True,
    "actions/curiosidades_loader.py":  False,  # opcional
    "actions/interaccion_loader.py":   False,  # opcional
    "actions/cuentos_loader.py":       True,
    "actions/corpus/palabras.xlsx":    True,
    "actions/corpus/cuentos.xlsx":     True,
    "actions/corpus/curiosidades.json": False,  # opcional
    "actions/corpus/frases_conversacionales.xlsx": False,  # opcional
    "domain.yml":                      True,
    "data/nlu.yml":                    True,
    "data/rules.yml":                  True,
    "data/stories.yml":                True,
    "server.py":                       True,
    "index.html":                      True,
}

for arch, obligatorio in archivos.items():
    p = HERE / arch
    if p.exists():
        ok(f"{arch} ({p.stat().st_size} bytes)")
    else:
        if obligatorio:
            err(f"FALTA: {arch}")
            errores += 1
        else:
            warn(f"opcional ausente: {arch}")

# ── 2. CONTENIDO DEL EXCEL DE PALABRAS ──────────────────────────────────
head("2. Contenido de palabras.xlsx")
try:
    from openpyxl import load_workbook
    xlsx = HERE / "actions" / "corpus" / "palabras.xlsx"
    wb = load_workbook(str(xlsx), read_only=True, data_only=True)
    ws = wb.active
    prefijos = {}
    num_con_no_hay = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        id_ = str(row[0]).strip()
        shp = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ""
        prefix = id_.split("_", 1)[0] if "_" in id_ else "?"
        if shp == "no hay":
            num_con_no_hay += 1
            continue
        prefijos[prefix] = prefijos.get(prefix, 0) + 1
    wb.close()

    print(f"     Prefijos encontrados: {dict(prefijos)}")
    print(f"     Filas con 'no hay' (se filtran): {num_con_no_hay}")

    if "num" in prefijos:
        ok(f"Categoría 'números' presente con {prefijos['num']} palabras")
    else:
        err("Categoría 'números' NO está en el Excel — usá el nuevo palabras.xlsx")
        errores += 1
except Exception as e:
    err(f"Error leyendo palabras.xlsx: {e}")
    errores += 1

# ── 3. CORPUS_LOADER LO LEE BIEN ────────────────────────────────────────
head("3. corpus_loader.py reconoce la nueva categoría")
sys.path.insert(0, str(HERE / "actions"))
try:
    import corpus_loader as cl
    cats = cl.categorias()
    print(f"     Categorías cargadas: {cats}")
    if "números" in cats:
        ok(f"corpus_loader incluye 'números' con {len(cl.palabras_de('números'))} palabras")
    else:
        err("corpus_loader NO incluye 'números' — actualizá corpus_loader.py")
        errores += 1
    totales = cl.total_por_categoria()
    print(f"     total_por_categoria(): {totales}")
except Exception as e:
    err(f"Error importando corpus_loader: {e}")
    errores += 1

# ── 4. DOMAIN.YML INCLUYE NÚMEROS ──────────────────────────────────────
head("4. domain.yml tiene 'números' en slot categoria_actual")
try:
    domain_text = (HERE / "domain.yml").read_text(encoding='utf-8')
    if "- números" in domain_text:
        ok("'números' presente en el slot categoria_actual")
    else:
        err("'números' NO está en domain.yml. Actualizá domain.yml y reentrená.")
        errores += 1
except Exception as e:
    err(f"Error leyendo domain.yml: {e}")
    errores += 1

# ── 5. NLU.YML INCLUYE EJEMPLOS DE NÚMEROS ────────────────────────────
head("5. nlu.yml tiene ejemplos para 'números'")
try:
    nlu_text = (HERE / "data" / "nlu.yml").read_text(encoding='utf-8')
    import re
    ejemplos = re.findall(r"^\s*-\s+.*números", nlu_text, re.MULTILINE | re.IGNORECASE)
    if len(ejemplos) >= 3:
        ok(f"NLU tiene {len(ejemplos)} ejemplos con 'números'")
    else:
        warn(f"Solo {len(ejemplos)} ejemplos con 'números' en NLU "
             "(el bot puede no reconocerlos bien)")
except Exception as e:
    err(f"Error leyendo nlu.yml: {e}")

# ── 6. SERVER.PY USA TOTAL DINÁMICO ───────────────────────────────────
head("6. server.py tiene la lógica dinámica de TOTAL_PALABRAS")
try:
    server_text = (HERE / "server.py").read_text(encoding='utf-8')
    if "_calcular_total_palabras" in server_text:
        ok("server.py incluye función _calcular_total_palabras")
        if "números" in server_text:
            ok("server.py reconoce 'números'")
        else:
            err("server.py no menciona 'números' — usá la versión nueva")
            errores += 1
    else:
        err("server.py NO tiene la lógica dinámica — usá la versión nueva")
        errores += 1
except Exception as e:
    err(f"Error: {e}")
    errores += 1

# ── 7. INDEX.HTML SIN ARGENTINISMOS ───────────────────────────────────
head("7. index.html sin argentinismos")
try:
    html_text = (HERE / "index.html").read_text(encoding='utf-8')
    import re
    patrones_arg = [r"\bquerés\b", r"\bpodés\b", r"\btenés\b", r"\bvos\b",
                    r"\bsumergite\b", r"\belegí\b", r"\bvení\b", r"\bandá\b",
                    r"\bdecime\b", r"\bvolvé\b", r"\busá\b"]
    detectados = []
    for p in patrones_arg:
        if re.search(p, html_text, re.IGNORECASE):
            detectados.append(p.replace(r"\b", ""))
    if not detectados:
        ok("Ningún argentinismo detectado en index.html")
    else:
        err(f"Argentinismos en index.html: {detectados}")
        errores += 1
    # CSS de tarjetas correcto?
    if ".categoria-btn .emoji" in html_text:
        ok("CSS de tarjeta de categoría con espacios correctos")
    elif ".categoria-btn.emoji" in html_text:
        err("CSS de tarjeta CORRUPTO: '.categoria-btn.emoji' sin espacio")
        errores += 1
except Exception as e:
    err(f"Error: {e}")

# ── 8. ACTIONS.PY SIN ARGENTINISMOS ───────────────────────────────────
head("8. actions.py sin argentinismos")
try:
    act = (HERE / "actions" / "actions.py").read_text(encoding='utf-8')
    detectados = []
    for p in patrones_arg:
        if re.search(p, act, re.IGNORECASE):
            detectados.append(p.replace(r"\b", ""))
    if not detectados:
        ok("Ningún argentinismo detectado en actions.py")
    else:
        err(f"Argentinismos en actions.py: {detectados}")
        errores += 1
except Exception as e:
    err(f"Error: {e}")

# ── 9. ENDPOINT /PROGRESO/ FUNCIONA Y DEVUELVE NÚMEROS ────────────────
head("9. Endpoint /progreso/ devuelve 'números' (requiere server.py corriendo)")
try:
    import urllib.request
    test_uuid = "test-diagnostico-" + str(os.getpid())
    url = f"http://localhost:8080/progreso/{test_uuid}"
    req = urllib.request.Request(url)
    with urllib.request.urlopen(req, timeout=3) as resp:
        data = json.loads(resp.read())
    cats_devueltas = [c["categoria"] for c in data.get("categorias", [])]
    print(f"     Categorías que devuelve el endpoint: {cats_devueltas}")
    if "números" in cats_devueltas:
        ok("Endpoint devuelve 'números' correctamente")
    else:
        err("Endpoint NO devuelve 'números' — REINICIÁ server.py")
        errores += 1
except (urllib.error.URLError, ConnectionRefusedError):
    warn("server.py no está corriendo. Iniciá: python3 server.py")
except Exception as e:
    warn(f"No se pudo verificar el endpoint: {e}")

# ── RESUMEN FINAL ─────────────────────────────────────────────────────
print(f"\n{C.BOLD}═══════════════════════════════════════════════════════════════")
if errores == 0:
    print(f"{C.OK}  TODO EN ORDEN — el sistema está correctamente configurado{C.END}")
else:
    print(f"{C.ERR}  {errores} PROBLEMA(S) DETECTADO(S) — leé arriba para los detalles{C.END}")
print(f"{C.BOLD}═══════════════════════════════════════════════════════════════{C.END}")
