"""
server.py — Servidor proxy para el frontend de Pishico Bot.

Endpoints:
  GET  /                          → sirve index.html
  GET  /status                    → estado de Rasa
  GET  /progreso/<sender_id>      → progreso del usuario desde la DB (JSON)
  GET  /cuentos/<sender_id>       → catálogo de cuentos + progreso (JSON)
  POST /login                     → valida código de acceso y activa al usuario
  POST /webhooks/rest/webhook     → proxy a Rasa

Uso:
    python3 server.py
"""

import json
import mimetypes
import sqlite3
import sys
import urllib.request
import urllib.error
from http.server import HTTPServer, BaseHTTPRequestHandler
from pathlib import Path

# ── Import resiliente del módulo db.py de actions/ ───────────────────────────
# Se usa para validar y activar códigos de acceso en el endpoint /login.
# Si falla la importación, el servidor sigue funcionando pero /login responde
# con error (útil en entornos donde no se quiere autenticación, ej. dev local
# sin migración a códigos).
_actions_path = Path(__file__).parent / "actions"
if str(_actions_path) not in sys.path:
    sys.path.insert(0, str(_actions_path))

try:
    from db import (
        validar_codigo, activar_codigo, info_usuario,
        iniciar_sesion, actualizar_actividad_sesion,
    )
    _DB_DISPONIBLE = True
except Exception as _db_err:
    print(f"[proxy] WARN: db.py no disponible ({_db_err}); /login no funcionará")
    _DB_DISPONIBLE = False
    def validar_codigo(c): return False
    def activar_codigo(c, nombre=None): return False
    def info_usuario(c): return None
    def iniciar_sesion(c): return None
    def actualizar_actividad_sesion(c): return False

RASA_WEBHOOK = "http://localhost:5005/webhooks/rest/webhook"
RASA_STATUS  = "http://localhost:5005/status"
PORT         = 8080

# El proxy puede estar fuera de la carpeta actions/, así que probamos varias ubicaciones
HERE         = Path(__file__).parent
HTML_FILE    = HERE / "index.html"
IMAGES_DIR   = HERE / "images"
DB_CANDIDATES = [
    HERE / "actions" / "progress.db",
    HERE / "progress.db",
    HERE.parent / "actions" / "progress.db",
]

CUENTOS_XLSX_CANDIDATES = [
    HERE / "actions" / "corpus" / "cuentos.xlsx",
    HERE / "corpus" / "cuentos.xlsx",
    HERE.parent / "actions" / "corpus" / "cuentos.xlsx",
]

# Metadatos visuales por cuento (emoji + descripción corta).
# Si en el futuro se agregan más cuentos, el frontend usa 📖 por defecto.
CUENTO_META = {
    "motelo_tigre": {
        "emoji": "🐢",
        "descripcion": "Una fábula sobre cómo el ingenio vence a la fuerza.",
    },
    "paujil_fiesta": {
        "emoji": "🦃",
        "descripcion": "El paujil descubre por qué tiene plumas blancas en la cola.",
    },
    "matrimonio_shipibo": {
        "emoji": "🏡",
        "descripcion": "Las dos pruebas que el yerno debía superar antes de casarse.",
    },
    "anciano_camungo": {
    "emoji": "🧓",
    "descripcion": "Un relato breve sobre pesca, naturaleza y sabiduría.",
    },
    "la_pesca": {
        "emoji": "🎣",
        "descripcion": "Una historia sobre la pesca y la vida cotidiana.",
    },
}

# Emojis por categoría — sumar entradas al añadir categorías nuevas
EMOJI_CAT = {
    "naturaleza": "🌿", "animales": "🦜", "cuerpo": "🫀",
    "colores": "🎨", "objetos": "🏺", "números": "🔢",
}

# Mapeo prefijo → categoría (debe estar sincronizado con corpus_loader.py)
_CATEGORIA_MAP = {
    "nat": "naturaleza", "ani": "animales", "cuer": "cuerpo",
    "col": "colores",    "obj": "objetos", "num": "números",
}

# Marcadores que indican ausencia de equivalente shipibo (palabras filtradas)
_MARCADORES_VACIO = {"no hay", "no_hay", "n/a", "-", ""}

# Localizar palabras.xlsx (mismas ubicaciones candidatas que cuentos.xlsx)
PALABRAS_XLSX_CANDIDATES = [
    HERE / "actions" / "corpus" / "palabras.xlsx",
    HERE / "corpus" / "palabras.xlsx",
    HERE.parent / "actions" / "corpus" / "palabras.xlsx",
]


def _palabras_xlsx_path():
    for c in PALABRAS_XLSX_CANDIDATES:
        if c.exists():
            return c
    return None


def _calcular_total_palabras():
    """
    Lee palabras.xlsx y cuenta palabras válidas por categoría.
    Reemplaza la constante hardcoded para que reflejar el corpus real.
    Si el Excel no se encuentra, usa fallback estático.
    """
    fallback = {
        "naturaleza": 11, "animales": 13, "cuerpo": 12,
        "colores": 6, "objetos": 9, "números": 10,
    }
    path = _palabras_xlsx_path()
    if not path:
        return fallback
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        conteo = {cat: 0 for cat in _CATEGORIA_MAP.values()}
        for row in ws.iter_rows(min_row=2, values_only=True):
            id_ = str(row[0]).strip() if row and row[0] else ""
            if not id_:
                continue
            shp = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ""
            if shp in _MARCADORES_VACIO:
                continue  # filtrar "no hay"
            prefix = id_.split("_", 1)[0] if "_" in id_ else ""
            cat = _CATEGORIA_MAP.get(prefix)
            if cat:
                conteo[cat] += 1
        wb.close()
        return {k: v for k, v in conteo.items() if v > 0}
    except Exception:
        return fallback


# Total dinámico computado al iniciar el servidor
TOTAL_PALABRAS = _calcular_total_palabras()


def _db_path():
    for c in DB_CANDIDATES:
        if c.exists():
            return c
    return None


def _leer_progreso(sender_id: str):
    """
    Lee progreso por categoría y cuento usando db.py (RDS).
    Reemplaza la lectura directa de SQLite local (legacy).

    Incluye:
      - categorias: resumen por nivel pedagógico (compatibilidad legacy)
      - cuentos:    resumen de fragmentos completados
      - scoring:    desglose ponderado (global + por categoría) — HITO 4
    """
    if not _DB_DISPONIBLE:
        return {"error": "DB no disponible", "categorias": [], "cuentos": [], "scoring": None}

    try:
        # db.py ya tiene la lógica completa con queries portables a PostgreSQL
        from db import (
            get_resumen_categorias,
            get_resumen_cuento,
            get_resumen_scoring_completo,
        )

        categorias = get_resumen_categorias(sender_id)
        cuentos_brutos = get_resumen_cuento(sender_id)

        # Renombrar campo "completados" si viene como "correctas" en algunas funciones
        cuentos = []
        for c in cuentos_brutos:
            cuentos.append({
                "cuento_id":   c.get("cuento_id"),
                "completados": c.get("completados", 0),
                "niveles":     c.get("niveles", {"2": 0, "1": 0, "0": 0}),
            })

        # Nuevo: scoring ponderado (global + por categoría)
        try:
            scoring = get_resumen_scoring_completo(sender_id)
        except Exception as e_scoring:
            print(f"[proxy] WARN scoring: {e_scoring}")
            scoring = None

        return {
            "categorias": categorias,
            "cuentos":    cuentos,
            "scoring":    scoring,
        }
    except Exception as e:
        print(f"[proxy] ERROR _leer_progreso: {e}")
        return {"error": str(e), "categorias": [], "cuentos": [], "scoring": None}


def _cuentos_xlsx_path():
    for c in CUENTOS_XLSX_CANDIDATES:
        if c.exists():
            return c
    return None


def _leer_cuentos(sender_id: str):
    """
    Lee cuentos.xlsx + progreso del usuario en la DB y devuelve metadatos
    enriquecidos para la biblioteca de cuentos.
    """
    xlsx = _cuentos_xlsx_path()
    if not xlsx:
        return {"error": "cuentos.xlsx no encontrado", "cuentos": []}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl no instalado", "cuentos": []}

    wb = load_workbook(str(xlsx), read_only=True, data_only=True)
    ws = wb["cuentos"] if "cuentos" in wb.sheetnames else wb.active

    # Agrupar fragmentos por cuento_id
    cuentos_raw = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cuento_id     = str(row[0]).strip()
        cuento_titulo = str(row[1]).strip() if row[1] else cuento_id
        respuesta_esp = str(row[5]).strip() if len(row) > 5 and row[5] else ""

        if cuento_id not in cuentos_raw:
            cuentos_raw[cuento_id] = {
                "id":      cuento_id,
                "titulo":  cuento_titulo,
                "total":   0,
                "palabras": [],
            }
        cuentos_raw[cuento_id]["total"] += 1
        if respuesta_esp and respuesta_esp not in cuentos_raw[cuento_id]["palabras"]:
            cuentos_raw[cuento_id]["palabras"].append(respuesta_esp)
    wb.close()

    # Leer progreso por cuento desde RDS via db.py
    progreso = {}
    if _DB_DISPONIBLE:
        try:
            from db import get_resumen_cuento
            resumen = get_resumen_cuento(sender_id)
            progreso = {r["cuento_id"]: r["completados"] for r in resumen}
        except Exception as e:
            print(f"[proxy] WARN _leer_cuentos: {e}")

    # Combinar metadatos
    resultado = []
    for cid, info in cuentos_raw.items():
        meta = CUENTO_META.get(cid, {"emoji": "📖", "descripcion": ""})
        completados = progreso.get(cid, 0)
        total = info["total"]
        resultado.append({
            "id":           cid,
            "titulo":       info["titulo"],
            "emoji":        meta["emoji"],
            "descripcion":  meta["descripcion"],
            "palabras":     info["palabras"],
            "total":        total,
            "completados":  min(completados, total),
            "porcentaje":   round(min(completados, total) / total * 100) if total else 0,
        })
    return {"cuentos": resultado}


class PishicoProxy(BaseHTTPRequestHandler):

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            try:
                content = HTML_FILE.read_bytes()
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(content)
            except FileNotFoundError:
                self._error(404, "index.html no encontrado")

        elif self.path == "/status":
            try:
                with urllib.request.urlopen(RASA_STATUS, timeout=3) as r:
                    self._json(200, json.loads(r.read()))
            except Exception:
                self._json(503, {"error": "Rasa no disponible"})

        elif self.path.startswith("/images/"):
            name = self.path.split("/images/", 1)[1]
            path = (IMAGES_DIR / name).resolve()
            if not str(path).startswith(str(IMAGES_DIR.resolve())) or not path.is_file():
                self._error(404, "Imagen no encontrada")
                return
            content_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.end_headers()
            self.wfile.write(path.read_bytes())

        elif self.path.startswith("/progreso/"):
            sid = self.path.split("/progreso/", 1)[1]
            if not sid:
                self._json(400, {"error": "Falta sender_id"})
                return
            data = _leer_progreso(sid)
            self._json(200, data)

        elif self.path.startswith("/cuentos/"):
            sid = self.path.split("/cuentos/", 1)[1]
            if not sid:
                self._json(400, {"error": "Falta sender_id"})
                return
            data = _leer_cuentos(sid)
            self._json(200, data)

        else:
            self._error(404, "Ruta no encontrada")

    def do_POST(self):
        # ── Login: valida y activa código de acceso ─────────────────────────
        if self.path == "/login":
            length = int(self.headers.get("Content-Length", 0))
            try:
                body = self.rfile.read(length) if length else b"{}"
                data = json.loads(body)
            except Exception:
                self._json(400, {"ok": False, "error": "JSON inválido"})
                return

            codigo = (data.get("codigo") or "").strip()
            nombre = (data.get("nombre") or "").strip() or None

            if not codigo:
                self._json(400, {"ok": False, "error": "Falta el código de acceso"})
                return

            if not _DB_DISPONIBLE:
                self._json(503, {"ok": False, "error": "Base de datos no disponible"})
                return

            if not validar_codigo(codigo):
                self._json(404, {
                    "ok": False,
                    "error": "Código no reconocido. Verifica con tu investigador.",
                })
                return

            # Detectar si es primer acceso ANTES de activar (para reporte al frontend)
            info_previa = info_usuario(codigo) or {}
            primera_vez = info_previa.get("fecha_primer_acceso") is None

            if not activar_codigo(codigo, nombre=nombre):
                self._json(500, {"ok": False, "error": "Error al activar el código"})
                return

            # Iniciar tracking de sesión de uso (tabla `sesiones`).
            # Si tenía sesiones abiertas previas, las cierra automáticamente.
            # No bloquear el login si falla por algún motivo.
            try:
                iniciar_sesion(codigo)
            except Exception as e_sesion:
                print(f"[proxy] WARN iniciar_sesion({codigo}): {e_sesion}")

            info_actual = info_usuario(codigo) or {}
            self._json(200, {
                "ok": True,
                "codigo_acceso": info_actual.get("codigo_acceso"),
                "nombre":        info_actual.get("nombre"),
                "primera_vez":   primera_vez,
            })
            return

        # ── Webhook Rasa: proxy transparente ───────────────────────────────
        if "/webhooks/rest/webhook" not in self.path:
            self._error(404, "Ruta no encontrada")
            return

        length = int(self.headers.get("Content-Length", 0))
        body   = self.rfile.read(length) if length else b""

        # Tracking de actividad: extraer sender_id del body y actualizar
        # la sesión abierta correspondiente. No bloquear el proxy si falla.
        if _DB_DISPONIBLE and body:
            try:
                body_json = json.loads(body)
                sender_id = (body_json.get("sender") or "").strip()
                if sender_id:
                    actualizar_actividad_sesion(sender_id)
            except Exception as e_track:
                # Logueo silencioso para no inundar la consola
                pass

        try:
            req = urllib.request.Request(
                RASA_WEBHOOK,
                data=body,
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=20) as resp:
                result = resp.read()
            self._json_raw(200, result)
        except urllib.error.URLError:
            self._json(503, [{"text": (
                "⚠️ No puedo conectar con Rasa. Asegúrate de que esté corriendo."
            )}])
        except Exception as e:
            self._json(500, [{"text": f"Error interno: {e}"}])

    # ── Helpers ───────────────────────────────────────────────────────
    def _json(self, code, data):
        self._json_raw(code, json.dumps(data, ensure_ascii=False).encode())

    def _json_raw(self, code, raw):
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(raw)

    def _error(self, code, msg):
        self.send_response(code)
        self.end_headers()
        self.wfile.write(msg.encode())

    def log_message(self, fmt, *args):
        if any(c in str(args) for c in ("500", "503")):
            print(f"[proxy] {self.address_string()} {args}")


def main():
    if not HTML_FILE.exists():
        print(f"⚠️  No se encontró {HTML_FILE}")
        return

    # Mostrar la URL de la BD que está usando db.py (RDS o SQLite según .env)
    db_info = "(no disponible)"
    if _DB_DISPONIBLE:
        try:
            from db import DATABASE_URL
            # Ocultar la password en la salida (formato user:pwd@host)
            if "@" in DATABASE_URL:
                db_info = DATABASE_URL.split("@", 1)[1]
            else:
                db_info = DATABASE_URL
        except Exception:
            db_info = "(error leyendo URL)"

    print("=" * 52)
    print("  Pishico Bot — Servidor frontend")
    print("=" * 52)
    print(f"  Frontend : http://0.0.0.0:{PORT}")
    print(f"  Rasa API : {RASA_WEBHOOK}")
    print(f"  DB       : {db_info}")
    print("=" * 52)

    server = HTTPServer(("0.0.0.0", PORT), PishicoProxy)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServidor detenido.")


if __name__ == "__main__":
    main()